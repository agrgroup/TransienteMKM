"""
Data extraction with proper negative barrier handling.

This version implements your assignment logic and adds safety clamping.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import logging
from openpyxl import load_workbook
import xlwings as xw
import math

logger = logging.getLogger(__name__)

class ExcelDataProcessor:
    """Handles Excel file operations with proper barrier assignment."""

    def __init__(self, excel_path: str):
        """Initialize with path to Excel file."""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

    def modify_local_environment(self, pH: float, V: float, output_path: str = None) -> str:
        """
        Modify pH and V values in the Local Environment sheet.
        
        Args:
            pH: New pH value
            V: New potential value
            output_path: Path for modified file (default: input_data.xlsx)
        
        Returns:
            Path to modified file
        """
        if output_path is None:
            output_path = "input_data.xlsx"

        try:
            workbook = load_workbook(filename=str(self.excel_path))
            sheet = workbook['Local Environment']

            # Update pH values
            self._update_column_values(sheet, 'pH', pH)
            # Update V values
            self._update_column_values(sheet, 'V', V)

            workbook.save(filename=output_path)
            logger.info(f"Excel modified successfully: pH={pH}, V={V}V → {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error modifying Excel file: {e}")
            raise

    def _update_column_values(self, sheet, column_name: str, new_value: float) -> None:
        """Update all values in a specific column."""
        column_index = None
        
        # Find column index
        for cell in sheet[1]:
            if cell.value == column_name:
                column_index = cell.column
                break

        if column_index is None:
            logger.warning(f"Column '{column_name}' not found")
            return

        # Update all non-None values in the column
        for row in range(2, sheet.max_row + 1):
            current_value = sheet.cell(row=row, column=column_index).value
            if current_value is not None:
                sheet.cell(row=row, column=column_index).value = new_value

    def read_column_data(self, sheet_name: str, column_name: str) -> List[Any]:
        """
        Read all values from a specific column using xlwings.
        
        Args:
            sheet_name: Name of the Excel sheet
            column_name: Name of the column
        
        Returns:
            List of values from the column
        """
        try:
            with xw.Book(str(self.excel_path)) as wb:
                sheet = wb.sheets[sheet_name]
                header_row = sheet.range('1:1').value

                if column_name not in header_row:
                    logger.warning(f"Column '{column_name}' not found in sheet '{sheet_name}'")
                    return []

                column_index = header_row.index(column_name) + 1
                column_values = sheet.range((2, column_index),
                                          (sheet.cells.last_cell.row, column_index)).value

                # Filter out None values and convert single values to list
                if not isinstance(column_values, list):
                    column_values = [column_values] if column_values is not None else []

                return [value for value in column_values if value is not None]

        except Exception as e:
            logger.error(f"Error reading column data: {e}")
            return []

    def _assign_barriers_with_safety(self, Ea: List[float], Eb: List[float], 
                                    delE_list: List[float]) -> Tuple[List[float], List[float]]:
        """
        Apply your assignment logic with safety clamping to prevent negative barriers.
        
        Assignment Logic:
        - Ea < 0: Ea = 0, Eb = -DelG_rxn
        - Eb < 0: Eb = 0, Ea = DelG_rxn  
        - Both Ea,Eb = 0: Use DelG_rxn (if DelG>0: Ea=DelG, Eb=0; else: Ea=0, Eb=-DelG)
        - Both ≥ 0: Use as extracted
        
        Safety: All final values are clamped to ≥ 0
        """
        Ea_final = Ea.copy()
        Eb_final = Eb.copy()
        delE_index = 0

        for i in range(len(Ea_final)):
            # Handle None/NaN by treating them as 0 for comparison
            ea_val = Ea_final[i] if Ea_final[i] is not None and not math.isnan(Ea_final[i]) else 0
            eb_val = Eb_final[i] if Eb_final[i] is not None and not math.isnan(Eb_final[i]) else 0

            # Get DelG_rxn value if available
            delE = None
            if delE_index < len(delE_list):
                delE = delE_list[delE_index]
                if delE is not None and not math.isnan(delE):
                    delE_index += 1
                else:
                    delE = None

            # Apply assignment logic
            if ea_val == 0 and eb_val == 0:
                # Case: Both barriers are zero - use DelG_rxn
                if delE is not None:
                    if delE > 0:
                        Ea_final[i] = delE
                        Eb_final[i] = 0.0
                        logger.debug(f"Rxn {i+1}: Both=0, DelG>0 → Ea={delE:.1f}, Eb=0")
                    else:
                        Ea_final[i] = 0.0
                        Eb_final[i] = -delE
                        logger.debug(f"Rxn {i+1}: Both=0, DelG≤0 → Ea=0, Eb={-delE:.1f}")
                else:
                    logger.warning(f"Rxn {i+1}: Both barriers zero but no DelG value available")
            
            elif ea_val < 0:
                # Case: Ea < 0 → Ea = 0, Eb = -DelG_rxn
                if delE is not None:
                    Ea_final[i] = 0.0
                    Eb_final[i] = -delE
                    logger.debug(f"Rxn {i+1}: Ea<0 → Ea=0, Eb={-delE:.1f}")
                else:
                    Ea_final[i] = 0.0  # Safety fallback
                    
            elif eb_val < 0:
                # Case: Eb < 0 → Eb = 0, Ea = DelG_rxn
                if delE is not None:
                    Ea_final[i] = delE
                    Eb_final[i] = 0.0
                    logger.debug(f"Rxn {i+1}: Eb<0 → Ea={delE:.1f}, Eb=0")
                else:
                    Eb_final[i] = 0.0  # Safety fallback
            
            # Final safety clamping - CRITICAL for preventing negative barriers
            Ea_final[i] = max(0.0, Ea_final[i])
            Eb_final[i] = max(0.0, Eb_final[i])
            
            # Log any clamping that occurred
            if Ea_final[i] == 0.0 and ea_val != 0:
                logger.warning(f"Rxn {i+1}: Ea clamped to 0 (was {ea_val:.1f})")
            if Eb_final[i] == 0.0 and eb_val != 0:
                logger.warning(f"Rxn {i+1}: Eb clamped to 0 (was {eb_val:.1f})")

        return Ea_final, Eb_final

    def extract_all_data(self, modified_excel_path: str) -> Dict[str, Any]:
        """
        Extract all necessary data from Excel file with proper barrier handling.
        
        Args:
            modified_excel_path: Path to the modified Excel file
        
        Returns:
            Dictionary containing all extracted data
        """
        try:
            # Read data using pandas for reliability
            df_reactions = pd.read_excel(modified_excel_path, sheet_name='Reactions')
            df_local_env = pd.read_excel(modified_excel_path, sheet_name='Local Environment')
            df_species = pd.read_excel(modified_excel_path, sheet_name='Input-Output Species')

            # Extract reaction data
            reactions = df_reactions['Reactions'].tolist()
            Ea_raw = self.read_column_data('Reactions', 'G_f') or df_reactions['G_f'].tolist()
            Eb_raw = self.read_column_data('Reactions', 'G_b') or df_reactions['G_b'].tolist()
            delE_list = self.read_column_data('Reactions', 'DelG_rxn') or df_reactions['DelG_rxn'].tolist()

            # *** APPLY YOUR ASSIGNMENT LOGIC WITH SAFETY CLAMPING ***
            Ea, Eb = self._assign_barriers_with_safety(Ea_raw, Eb_raw, delE_list)
            
            # Validate no negative barriers remain
            negative_count = sum(1 for i in range(len(Ea)) if Ea[i] < 0 or Eb[i] < 0)
            if negative_count > 0:
                logger.error(f"❌ {negative_count} negative barriers still present after assignment!")
            else:
                logger.info(f"✅ All {len(Ea)} activation barriers are non-negative")

            # Extract environment data
            V = df_local_env['V'].iloc[0]
            pH = df_local_env['pH'].iloc[0]
            P = df_local_env['Pressure'].iloc[0]

            # Extract species data
            gases = df_species['Species'].tolist()
            concentrations = (self.read_column_data('Input-Output Species', 'Input MKMCXX') or
                             df_species.get('Input MKMCXX', []).tolist())

            # Parse reactions
            parsed_reactions = self._parse_reactions(reactions)

            # Find adsorbates
            adsorbates = self._extract_adsorbates(parsed_reactions)

            return {
                'reactions': reactions,
                'Ea': Ea,
                'Eb': Eb,
                'V': V,
                'pH': pH,
                'P': P,
                'gases': gases,
                'concentrations': concentrations,
                'adsorbates': adsorbates,
                'activity': np.zeros(len(adsorbates)),
                **parsed_reactions
            }

        except Exception as e:
            logger.error(f"Error extracting data: {e}")
            raise

    def _parse_reactions(self, reactions: List[str]) -> Dict[str, List[str]]:
        """Parse reaction strings into reactants and products."""
        reactant1, reactant2, reactant3 = [], [], []
        product1, product2, product3 = [], [], []

        for rxn in reactions:
            try:
                # Split reaction into reactants and products
                reactants_str, products_str = rxn.strip().split("→")

                # Parse reactants
                reactants = [r.strip() for r in reactants_str.split("+")]
                reactant1.append(f"{{{reactants[0]}}}")
                reactant2.append(f"{{{reactants[1]}}}" if len(reactants) > 1 else "")
                reactant3.append(f"{{{reactants[2]}}}" if len(reactants) > 2 else "")

                # Parse products
                products = [p.strip() for p in products_str.split("+")]
                product1.append(f"{{{products[0]}}}")
                product2.append(f"{{{products[1]}}}" if len(products) > 1 else "")
                product3.append(f"{{{products[2]}}}" if len(products) > 2 else "")

            except Exception as e:
                logger.error(f"Error parsing reaction '{rxn}': {e}")
                # Add empty entries to maintain list consistency
                for lst in [reactant1, reactant2, reactant3, product1, product2, product3]:
                    lst.append("")

        return {
            'Reactant1': reactant1,
            'Reactant2': reactant2,
            'Reactant3': reactant3,
            'Product1': product1,
            'Product2': product2,
            'Product3': product3
        }

    def _extract_adsorbates(self, parsed_reactions: Dict[str, List[str]]) -> List[str]:
        """Extract unique adsorbates from parsed reactions."""
        adsorbates = set()

        # Check all reactants and products for adsorbates (containing *)
        for key in ['Reactant1', 'Reactant2', 'Product1', 'Product2']:
            for item in parsed_reactions.get(key, []):
                if "*" in item and item != "":
                    species = item.strip("{}").strip()
                    if species != "*":
                        adsorbates.add(species)

        return list(adsorbates)


def data_extract(pH: float, V: float, inp_path: str) -> Tuple:
    """
    Main data extraction function with proper negative barrier handling.
    
    Returns:
        Tuple containing all extracted data in original order
    """
    processor = ExcelDataProcessor(inp_path)

    # Modify Excel file with new pH and V values
    modified_path = processor.modify_local_environment(pH, V)

    # Extract all data
    data = processor.extract_all_data(modified_path)

    return (
        data['gases'],
        data['concentrations'],
        data['adsorbates'],
        data['activity'],
        data['Reactant1'],
        data['Reactant2'],
        data['Reactant3'],
        data['Product1'],
        data['Product2'],
        data['Product3'],
        data['Ea'],
        data['Eb'],
        data['P'],
        data['reactions']
    )