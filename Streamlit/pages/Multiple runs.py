import streamlit as st
import pandas as pd
import os
import sys
import subprocess
import xlwings as xw
import traceback
import shutil
import pythoncom
import numpy as np

from openpyxl import load_workbook

# Dynamically resolve base directory
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.abspath(os.path.join(BASE_DIR, '..')))

# Import custom modules
from dependencies import *
from inp_file_gen import *
from inp_file import *
from mkm_parameters import *

st.set_page_config(page_title="MKM Input File Generator and Solver", page_icon=":coffee:")
st.title('☕ MKM Input File Generator and Solver for Multiple Runs')
from io import StringIO
st.page_icon = ":coffee:"

# Save the initial working directory
def main():
    # Dynamically reset to the main working directory
    os.chdir(BASE_DIR)

    uploaded_file = st.file_uploader("Upload Excel File", type="xlsx")
    if not uploaded_file:
        st.warning("Please upload an Excel file to continue.")
        return

    pH_l = [round(x * 0.5, 1) for x in range(29)]  # 0.0 to 14.0
    V_l = [round(x * 0.1 - 1.0, 1) for x in range(21)]  # -1.0 to 1.0

    selected_pHs = st.multiselect("Select pH Values", pH_l)
    st.write(f"Selected pH values: {selected_pHs}")

    selected_Vs = st.multiselect("Select Potential Values", V_l)
    st.write(f"Selected Potential values: {selected_Vs}")

    if not selected_pHs or not selected_Vs:
        st.warning("Please select both pH and V values to proceed.")
        return

    for pH in selected_pHs:
        parent_folder = os.path.join(BASE_DIR, f"pH_{pH}")
        os.makedirs(parent_folder, exist_ok=True)

        for V in selected_Vs:
            child_folder = os.path.join(parent_folder, f"V_{V}")
            os.makedirs(child_folder, exist_ok=True)

            try:
                # Save uploaded file into this child directory
                excel_path = os.path.join(child_folder, "inp_file.xlsx")
                with open(excel_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                # Modify Excel file
                modify_excel(child_folder, pH, V)
                st.write(f"Modified Excel written to: {child_folder}")

                # Load necessary sheets
                data1 = pd.read_excel(excel_path, sheet_name="Reactions")
                data2 = pd.read_excel(excel_path, sheet_name="Local Environment")
                data3 = pd.read_excel(excel_path, sheet_name="Input-Output Species")

                # Extract input data
                global gases, rxn, concentrations, Ea, Eb, P
                rxn = data1["Reactions"]
                Ea = read_formulas(excel_path, "Reactions", "G_f")["G_f"]
                Eb = read_formulas(excel_path, "Reactions", "G_b")["G_b"]
                concentrations = read_formulas(excel_path, "Input-Output Species", "Input MKMCXX")["Input MKMCXX"]
                gases = data3["Species"].tolist()

                V = data2["V"][0]
                pH = data2["pH"][0]
                P = data2["Pressure"][0]

                # Intermediates
                global adsorbates, activity, Reactant1, Reactant2, Reactant3, Product1, Product2, Product3, Time, Temp, Abstol, Reltol
                Reactant1, Reactant2, Reactant3 = [], [], []
                Product1, Product2, Product3 = [], [], []
                adsorbates = []

                for r in rxn:
                    lhs, rhs = r.strip().split("→")
                    lhs_parts = [x.strip() for x in lhs.split("+")]
                    rhs_parts = [x.strip() for x in rhs.split("+")]

                    Reactant1.append(f"{{{lhs_parts[0]}}}")
                    Reactant2.append(f"{{{lhs_parts[1]}}}" if len(lhs_parts) > 1 else "")
                    Reactant3.append(f"{{{lhs_parts[2]}}}" if len(lhs_parts) > 2 else "")

                    Product1.append(f"{{{rhs_parts[0]}}}")
                    Product2.append(f"{{{rhs_parts[1]}}}" if len(rhs_parts) > 1 else "")
                    Product3.append(f"{{{rhs_parts[2]}}}" if len(rhs_parts) > 2 else "")

                for idx in Reactant1 + Reactant2 + Product1 + Product2:
                    species = idx.strip("{}")
                    if "*" in species and species not in adsorbates and species != "*":
                        adsorbates.append(species)

                activity = np.zeros(len(adsorbates))

                os.chdir(child_folder)
                inp_file_gen(
                    rxn, pH, V, gases, concentrations, adsorbates, activity,
                    Reactant1, Reactant2, Reactant3, Product1, Product2, Product3,
                    Ea, Eb, P, Temp, Time, Abstol, Reltol
                )
                st.success(f"Input file generated for pH={pH}, V={V}")

                input_file_path = os.path.join(child_folder, "input_file.mkm")  # <- add this line here


                # if st.button("Run Solver"):
                #     result_message, success = run_executable(input_file_path)
                #     if success:
                #         st.success(result_message)
                #         coverage() 
                #     else:
                #         st.error(result_message)   

                if st.button("Run Solver", key=f"run_solver_{pH}_{V}"):
                    result_message, success = run_executable(input_file_path)
                    if success:
                        st.success(result_message)
                        coverage()
                    else:
                        st.error(result_message)        

                os.chdir(BASE_DIR)  # Reset for next run

            except Exception as e:
                st.error(f"Failed to process pH={pH}, V={V}: {e}")
                st.text(traceback.format_exc())

                        
    
from openpyxl import load_workbook

def modify_excel(children_folder, pH,potential):
    try:
        # Load the workbook
        workbook = load_workbook(filename="input.xlsx")
        
        # Check if the sheet exists
        if "Local Environment" not in workbook.sheetnames:
            st.error("Sheet 'Local Environment' not found.")
            return
        
        # Access the sheet
        sheet = workbook["Local Environment"]

        # Modify B2 (Potential) and C2 (pH)
        sheet["B2"].value = potential
        sheet["C2"].value = pH

        # Save the workbook
        workbook.save(filename="inp_file.xlsx")
        shutil.copy("inp_file.xlsx", children_folder)
        st.success("Workbook updated and saved as 'inp_file.xlsx'.")
    except FileNotFoundError:
        st.error("The file 'Inp_file.xlsx' was not found. Please upload it.")
    except PermissionError:
        st.error("Permission error: Ensure the file is not open in another program.")
    except Exception as e:
        st.error(f"An unexpected error occurred: {e}")

                     

def read_formulas(file_name, sheet_name, column_name):
    """
    Recalculates the formulas in an Excel file, saves the file, and returns the dataframe with the updated values.
    
    Args:
    file_name (str): Path to the Excel file.
    sheet_name (str): Name of the sheet to process.
    column_name (str): The column name to focus on in the resulting DataFrame.
    
    Returns:
    pd.DataFrame: DataFrame with the computed values from the specified column.
    """
    try:
        # Initialize COM thread for Excel (important for multi-threaded environments like Streamlit)
        pythoncom.CoInitialize()
        # Open Excel file
        app = xw.App(visible=False)  # Open Excel in the background
        wb = xw.Book(file_name)
        sheet = wb.sheets[sheet_name]

        # Recalculate all formulas in the workbook
        wb.app.calculation = 'automatic'
        wb.save(file_name)  # Save the file after recalculation
        wb.close()
        app.quit()

        # Read the recalculated file with pandas
        df = pd.read_excel(file_name, engine="openpyxl", sheet_name=sheet_name)
        
        # Optionally, focus on the specified column if needed
        if column_name in df.columns:
            return df[[column_name]]  # Return only the specified column
        else:
            raise ValueError(f"Column '{column_name}' not found in the sheet.")
    
    except Exception as e:
        st.write(f"Error: {e}")
        print(f"Error: {e}")
        return None 

def run_executable(input_file):
    #executable_path = os.path.join(os.getcwd(),'/bin/mkmcxx')    
    # Get the directory of the current script
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct relative path to the executable one directory up, inside "bin"
    executable_path = os.path.join(current_dir, "..", "bin", "mkmcxx.exe")

    # Normalize the path (resolves ../ properly)
    executable_path = os.path.normpath(executable_path)
    st.write(executable_path)
    if os.path.exists(executable_path):
        st.write(f"Executable found at: {executable_path}")
        try:
            # Run the executable with the input file using subprocess
            result = subprocess.run([executable_path, '-i', input_file],
                                    capture_output=True, text=True)
            
            # Print the stdout (standard output) and stderr (standard error)
            st.write("Solver Output (stdout):")
            st.text(result.stdout)  # This will show the standard output in Streamlit

            if result.stderr:
                st.write("Solver Error Output (stderr):")
                st.text(result.stderr)  # This will show the error output (if any)
            
            # Check for success or failure
            if result.returncode == 0:
                return "Solver ran successfully!", True
            else:
                return f"Error running solver: {result.stderr}", False
        except Exception as e:
            return f"Error executing command: {str(e)}", False
    else:
        return "Executable not found at the given path.", False
    


def get_val (cov_path):   
    cov_file = open(cov_path)
    
    cov_val = []
    
    cov_lines = cov_file.readlines()
    
    adsorbate_keys = cov_lines[0].strip().split()
    
    cov_dat_dict = {}
    
    for key in adsorbate_keys:
        cov_dat_dict[key] = []
    
    for line in cov_lines[1:]:
        vals = line.strip().split()
        vals = list(map(lambda x : float(x), vals))
        c=0
        for key in adsorbate_keys:
            cov_dat_dict[key].append(vals[c])
            c+=1
            
    cov_file.close()
    return cov_dat_dict
    
def coverage():
    coverage_file_path = "run/range/coverage.dat"
    
    if os.path.exists(coverage_file_path):
        covs = get_val(coverage_file_path)
        covs_relevant ={}

        for key in covs.keys():
            if '*' in key:
                covs_relevant[key] = covs[key]
        #Read the data from the coverage.dat file
        # list coverrage dictionary
        covs_relevant_df = pd.DataFrame(covs_relevant)
        covs_relevant_df = covs_relevant_df.T
        covs_relevant_df = covs_relevant_df.reset_index()
        covs_relevant_df.columns = ['Adsorbates', 'Coverage']
        st.write("Coverage Data:")
        st.write(covs_relevant_df)
    else:
        st.error("coverage.dat file not found in the expected directory.")     


if __name__ == "__main__":
    main()    